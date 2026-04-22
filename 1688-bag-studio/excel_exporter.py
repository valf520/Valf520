from __future__ import annotations

import io
from datetime import datetime
from pathlib import Path

import httpx
import pandas as pd
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font, PatternFill


COLUMNS = [
    "关键词",
    "商品标题",
    "详情网址",
    "图片",
    "图片URL",
    "适用性别",
    "箱包大小",
    "上市年份季节",
    "风格",
    "箱包潮流款式",
    "材质",
    "价格",
    "店铺名",
    "抓取时间",
]


def _download_image_bytes(url: str) -> bytes | None:
    if not url:
        return None
    try:
        with httpx.Client(timeout=10, follow_redirects=True) as client:
            resp = client.get(url)
            if resp.status_code == 200 and resp.content:
                return resp.content
    except httpx.HTTPError:
        return None
    return None


def export_to_excel(rows: list[dict], export_dir: Path) -> Path:
    export_dir.mkdir(exist_ok=True)
    file_path = export_dir / f"1688_箱包采集_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    normalized = []
    for row in rows:
        normalized.append(
            {
                "关键词": row.get("关键词", ""),
                "商品标题": row.get("商品标题", ""),
                "详情网址": row.get("详情网址", ""),
                "图片": "",
                "图片URL": row.get("图片URL", ""),
                "适用性别": row.get("适用性别", "未知"),
                "箱包大小": row.get("箱包大小", "未知"),
                "上市年份季节": row.get("上市年份季节", "未知"),
                "风格": row.get("风格", "未知"),
                "箱包潮流款式": row.get("箱包潮流款式", "未知"),
                "材质": row.get("材质", "未知"),
                "价格": row.get("价格", "未知"),
                "店铺名": row.get("店铺名", "未知"),
                "抓取时间": row.get("抓取时间", ""),
            }
        )

    df = pd.DataFrame(normalized, columns=COLUMNS)
    df.to_excel(file_path, index=False)

    wb = load_workbook(file_path)
    ws = wb.active
    ws.title = "1688箱包数据"

    header_fill = PatternFill(fill_type="solid", start_color="1777FF", end_color="1777FF")
    for cell in ws[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 36
    ws.column_dimensions["C"].width = 42
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 48
    for col in "FGHIJKLMN":
        ws.column_dimensions[col].width = 14

    for row_id in range(2, ws.max_row + 1):
        ws.row_dimensions[row_id].height = 82
        for col_id in range(1, ws.max_column + 1):
            ws.cell(row=row_id, column=col_id).alignment = Alignment(vertical="center", wrap_text=True)

        img_url = ws.cell(row=row_id, column=5).value
        image_bytes = _download_image_bytes(img_url)
        if image_bytes:
            try:
                stream = io.BytesIO(image_bytes)
                ximg = XLImage(stream)
                ximg.width = 72
                ximg.height = 72
                ws.add_image(ximg, f"D{row_id}")
            except Exception:
                ws.cell(row=row_id, column=4).value = "图片下载失败"
        else:
            ws.cell(row=row_id, column=4).value = "无图片"

    wb.save(file_path)
    return file_path

